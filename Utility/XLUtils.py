import openpyxl


def get_row_count(file,sheet_name):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet_name]
    return sheet.max_row


def get_col_count(file,sheet_name):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet_name]
    return sheet.max_column


def read_data(file,sheet_name,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet_name]
    return sheet.cell(row=rownum,column=colnum).value


def write_data(file,sheet_name,rownum,colnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet_name]
    sheet.cell(row=rownum,column=colnum).value=data
    workbook.save(file)

# path="D:\Coding\PageObjectSelenium\TestData\LoginData.xlsx"
# rows = get_row_count(path, 'Sheet1')
# print(rows)
# for r in range(2,rows + 1):
#     username = read_data(path, "Sheet1", r, 1)
#     print(username)